"""
exporter.py
===========
Exports clean and flagged records to a professionally formatted Excel workbook.

Workbook structure:
  Sheet 1 - "Data"    : Clean, validated records with frozen header row,
                         alternating row shading, and auto-width columns.
  Sheet 2 - "Flagged" : Invalid records, each annotated with a flag reason.
  Sheet 3 - "Summary" : Run metadata (date, counts, elapsed time, source).

Requires: pip install openpyxl
"""

import logging
from pathlib import Path
from typing import Any, Dict, List

log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── Output schema ─────────────────────────────────────────────────────
DATA_FIELDS: List[str] = [
    "Name", "Phone", "Website", "Postcode", "Category", "Source"
]
FLAG_FIELDS: List[str] = DATA_FIELDS + ["Flag Reason"]

# ── Styling constants ─────────────────────────────────────────────────
_HEADER_BG     = "1F4E79"   # dark navy
_HEADER_FG     = "FFFFFF"   # white text
_ALT_ROW_BG    = "DCE6F1"   # light blue stripe
_MAX_COL_WIDTH = 60


def _make_header_fill() -> "PatternFill":
    """Create a dark-navy header fill object."""
    return PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")


def _make_alt_fill() -> "PatternFill":
    """Create a light-blue alternating-row fill object."""
    return PatternFill(start_color=_ALT_ROW_BG, end_color=_ALT_ROW_BG, fill_type="solid")


def export_excel(
    clean_rows: List[Dict[str, str]],
    flagged_rows: List[Dict[str, str]],
    output_path: Path,
    stats: Dict[str, Any],
) -> None:
    """
    Write a formatted Excel workbook containing clean data, flagged data,
    and a run-summary sheet.

    Args:
        clean_rows   : Validated records (written to the "Data" sheet).
        flagged_rows : Records that failed validation (written to "Flagged").
        output_path  : Full path for the .xlsx output file.
        stats        : Run metadata dict written to the "Summary" sheet.

    Raises:
        ImportError: If openpyxl is not installed.
        OSError    : If the output directory cannot be created or the file
                     cannot be written.
    """
    if not _OPENPYXL_OK:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build fills once per export call and share across all sheets
    header_fill = _make_header_fill()
    alt_fill    = _make_alt_fill()

    wb = openpyxl.Workbook()

    # Sheet 1: Clean data
    ws_data = wb.active
    ws_data.title = "Data"
    _write_data_sheet(ws_data, clean_rows, DATA_FIELDS, header_fill, alt_fill)

    # Sheet 2: Flagged records
    ws_flagged = wb.create_sheet("Flagged")
    _write_data_sheet(ws_flagged, flagged_rows, FLAG_FIELDS, header_fill, alt_fill)

    # Sheet 3: Run summary
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, stats, header_fill)

    wb.save(output_path)
    log.info("Excel workbook written: %s", output_path)
    log.info(
        "  -> Data: %d rows | Flagged: %d rows",
        len(clean_rows), len(flagged_rows),
    )


# ── Private sheet builders ────────────────────────────────────────────

def _write_data_sheet(
    ws: "Worksheet",
    rows: List[Dict[str, str]],
    fields: List[str],
    header_fill: "PatternFill",
    alt_fill: "PatternFill",
) -> None:
    """
    Write field headers and data rows to a worksheet.

    Applies:
      - Bold white text on a dark navy header row.
      - Frozen header row (row 1).
      - Alternating light-blue shading on even data rows.
      - Auto-width columns (capped at _MAX_COL_WIDTH characters).
    """
    # Header row
    ws.append(fields)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color=_HEADER_FG)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # Data rows
    for i, row in enumerate(rows, start=2):
        ws.append([row.get(field, "") for field in fields])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = alt_fill

    _set_column_widths(ws, fields, rows)


def _write_summary_sheet(
    ws: "Worksheet",
    stats: Dict[str, Any],
    header_fill: "PatternFill",
) -> None:
    """
    Write a two-column summary table (Field | Value) with a formatted header.
    """
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_HEADER_FG)
        cell.fill = header_fill

    for key, value in stats.items():
        ws.append([key, str(value)])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36


def _set_column_widths(
    ws: "Worksheet",
    fields: List[str],
    rows: List[Dict[str, str]],
) -> None:
    """Set each column width to the longest value in that column (header or data)."""
    for col_idx, field in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(field)
        for row in rows:
            val = str(row.get(field, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 4, _MAX_COL_WIDTH)
